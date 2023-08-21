import datetime
import re
import sqlite3
import os
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import StatesGroup, State
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from openpyxl import Workbook
import random
from system.dispatcher import dp, bot
from openpyxl import Workbook
from openpyxl.styles import PatternFill

class RecordingQuestions(StatesGroup):
    """Расчет стоимости доставки"""
    application_date = State()  # Дата
    guest_number = State()  # Номер гостя
    name_of_the_guest = State()  # Имя гостя
    application_time = State()  # Время
    commentary = State()  # Комментарии


@dp.callback_query_handler(lambda c: c.data == 'apply')
async def application_record_handler(callback_query: types.CallbackQuery):
    """Ввод даты для отчета кнопки (Подать заявку)"""
    await bot.send_message(callback_query.from_user.id,
                           "Введите дату 📆 в формате День.Месяц (например, 25.07):")
    await RecordingQuestions.application_date.set()


@dp.message_handler(state=RecordingQuestions.application_date)
async def process_guest_number(message: types.Message, state: FSMContext):
    """Обработчик для ввода номера гостя"""
    date_pattern = r"\d{2}\.\d{2}"  # Регулярное выражение (dd.mm)
    if not re.match(date_pattern, message.text):
        await bot.send_message(message.chat.id, "Введите дату 📆 в формате День.Месяц (например, 25.07):")
        return
    # Преобразование пользовательского ввода в объект datetime
    try:
        # Добавьте текущий год к пользовательскому вводу, чтобы создать полный формат даты.
        current_year = datetime.date.today().year
        full_date_str = message.text + "." + str(current_year)
        application_date = datetime.datetime.strptime(full_date_str, "%d.%m.%Y").date()
    except ValueError:
        await bot.send_message(message.chat.id,
                               "Введите корректную дату 📆 в формате День.Месяц (например, 25.07):")
        return
    # Получить текущую дату
    current_date = datetime.date.today()
    if application_date < current_date:
        current_date_formatted = current_date.strftime("%d.%m")
        await bot.send_message(message.chat.id,
                               f"Нельзя выбрать дату, которая уже прошла текущая дата: {current_date_formatted}.\n\n Введите корректную дату 📆:")
        return

    await state.update_data(application_date=message.text)
    await bot.send_message(message.chat.id, "👥 Введите номер гостя:")
    await RecordingQuestions.guest_number.set()


@dp.message_handler(state=RecordingQuestions.guest_number)
async def process_guest_name(message: types.Message, state: FSMContext):
    """Обработчик для ввода Ф.И.О. гостя"""
    await state.update_data(guest_number=message.text)
    await bot.send_message(message.chat.id, "👤 Введите Ф.И.О. гостя:")
    await RecordingQuestions.name_of_the_guest.set()


@dp.message_handler(state=RecordingQuestions.name_of_the_guest)
async def process_application_time(message: types.Message, state: FSMContext):
    """Обработчик для ввода времени"""
    await state.update_data(name_of_the_guest=message.text)
    await bot.send_message(message.chat.id, "Введите время ⏳:")
    await RecordingQuestions.application_time.set()


@dp.message_handler(state=RecordingQuestions.application_time)
async def process_application_time(message: types.Message, state: FSMContext):
    # Просто сохраняем текст времени в хранилище, без дополнительной проверки
    application_time = message.text
    await state.update_data(application_time=application_time)
    await bot.send_message(message.chat.id, "💬 Введите комментарий:")
    await RecordingQuestions.commentary.set()


@dp.message_handler(state=RecordingQuestions.commentary)
async def process_price(message: types.Message, state: FSMContext):
    """Конечные данные отчета"""
    try:
        commentary = message.text  # Комментарий, введенный пользователем
        data = await state.get_data()  # Получаем данные из хранилища
        application_date = data.get('application_date')
        guest_number = data.get('guest_number')
        name_of_the_guest = data.get('name_of_the_guest')
        application_time = data.get('application_time')

        id_number = random.randint(100, 99999)  # Генерируем номер заявки случайным образом
        print(f"Сгенерирована заявка с номером {id_number}")

        data['commentary'] = commentary  # Добавление «комментария» в словарь данных
        data['id_number'] = id_number  # Добавление ID в словарь данных
        await state.update_data(data)  # Обновите данные в хранилище

        message_text = (f"<b>Данные:</b>\n\n"
                        f"<b>📂 ID:</b> <i>{id_number}</i>\n"
                        f"<b>📆 Дата:</b> <i>{application_date}</i>\n"
                        f"<b>👥 Номер гостя:</b> <i>{guest_number}</i>\n"
                        f"<b>👤 Ф.И.О. гостя:</b> <i>{name_of_the_guest}</i>\n"
                        f"<b>⏳ Время:</b> <i>{application_time}</i>\n"
                        f"<b>💬 Комментарий:</b> <i>{commentary}</i>\n"
                        "\n<b>Для возврата в начало нажмите</b> /start")
        # Создать встроенные кнопки "Отредактировать заявку" и "Редактировать заявку"
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(InlineKeyboardButton("Отправить заявку", callback_data="send_application"),
                   InlineKeyboardButton("Редактировать заявку", callback_data="edit_application"))
        await bot.send_message(message.chat.id, message_text, parse_mode='HTML', reply_markup=markup)
    except Exception as e:
        await bot.send_message(message.chat.id, "Произошла ошибка при обработке данных.")
        print(e)


@dp.callback_query_handler(lambda c: c.data == 'send_application', state=RecordingQuestions.commentary)
async def send_application_handler(callback_query: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()  # Получить сохраненные данные
        application_date = data.get('application_date')
        guest_number = data.get('guest_number')
        name_of_the_guest = data.get('name_of_the_guest')
        application_time = data.get('application_time')
        commentary = data.get('commentary')
        number = ""
        id_number = data.get('id_number')
        record_check = ""

        # Здесь вы можете использовать свое соединение SQLite и данные INSERT в таблицу
        # Замените 'database/database.db' именем вашего файла базы данных SQLite.
        conn = sqlite3.connect('database/database.db')
        cursor = conn.cursor()
        # Замените reports на имя вашей таблицы в базе данных.
        cursor.execute('''CREATE TABLE IF NOT EXISTS reports
                          (id_number, application_date, guest_number, number,  name_of_the_guest,
                           application_time, commentary, record_check)''')
        # Вставьте данные в таблицу
        cursor.execute("INSERT INTO reports VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                       (id_number, application_date, guest_number, number, name_of_the_guest, application_time,
                        commentary, record_check))
        conn.commit()
        conn.close()
        mes_text = ("<b>Ваша заявка была успешно отправлена! 🚀</b>"
                    "\n\nДля возврата в начало нажмите /start")
        await bot.send_message(callback_query.from_user.id, mes_text)
        await state.finish()
    except Exception as e:
        await bot.send_message(callback_query.from_user.id, "Произошла ошибка при отправке заявки.")
        print(e)


@dp.callback_query_handler(lambda c: c.data == 'edit_application', state='*')
async def edit_application_handler(callback_query: types.CallbackQuery, state: FSMContext):
    await state.finish()  # Сбросить состояние, чтобы начать сначала
    await application_record_handler(callback_query)  # Начните записывать вопросы с самого начала


@dp.callback_query_handler(lambda c: c.data == 'report_for_today')
async def send_report_for_today(callback_query: types.CallbackQuery):
    try:
        # Здесь вы можете использовать свое соединение SQLite для получения данных для современных приложений.
        # Замените 'database/database.db' именем вашего файла базы данных SQLite.
        conn = sqlite3.connect('database/database.db')
        cursor = conn.cursor()
        # Получить день и месяц сегодняшней даты
        today_day_month = datetime.date.today().strftime("%d.%m")
        # Замените reports на имя вашей таблицы в базе данных.
        cursor.execute("SELECT * FROM reports WHERE application_date LIKE ?", (f"%{today_day_month}%",))
        rows = cursor.fetchall()
        if not rows:
            await bot.send_message(callback_query.from_user.id, "На сегодня нет заявок.")
        else:
            workbook = Workbook()  # Создание книги и листа Excel
            sheet = workbook.active
            # Напишите строку заголовка
            header = ['ID заявки', 'Дата', 'Номер гостя', 'Номер стола', 'Ф.И.О. гостя', 'Время', 'Комментарий']
            sheet.append(header)

            # Создаем объект PatternFill для изменения цвета
            bad_record_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

            for row in rows:
                if row[7] == 'bad':
                    row_data = list(row)[:7]  # Берем только первые 7 элементов
                    sheet.append(row_data)
                    for cell in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row, min_col=1, max_col=7):
                        for col_idx in range(1, len(cell) + 1):
                            cell[col_idx - 1].fill = bad_record_fill
                else:
                    sheet.append(row[:7])
            # Сохраните файл Excel
            file_name = f"отчет_ЗА_СЕГОДНЯ_{datetime.date.today().strftime('%Y-%m')}.xlsx"
            workbook.save(file_name)
            # Отправить файл Excel пользователю
            with open(file_name, 'rb') as file:
                await bot.send_document(callback_query.from_user.id, file)
            os.remove(file_name)  # Удаление файла
        conn.close()
    except Exception as e:
        await bot.send_message(callback_query.from_user.id, "Произошла ошибка при получении отчета за сегодня.")
        print(e)


@dp.callback_query_handler(lambda c: c.data == 'report_for_tomorrow')
async def send_report_for_tomorrow(callback_query: types.CallbackQuery):
    try:
        # Здесь вы можете использовать свое соединение SQLite для получения данных для приложений завтрашнего дня.
        # Замените 'database/database.db' именем вашего файла базы данных SQLite.
        conn = sqlite3.connect('database/database.db')
        cursor = conn.cursor()
        # Получить день и месяц завтрашней даты
        tomorrow_day_month = (datetime.date.today() + datetime.timedelta(days=1)).strftime("%d.%m")
        # Замените reports на имя вашей таблицы в базе данных.
        cursor.execute("SELECT * FROM reports WHERE application_date LIKE ?", (f"%{tomorrow_day_month}%",))
        rows = cursor.fetchall()
        if not rows:
            await bot.send_message(callback_query.from_user.id, "На завтра нет заявок.")
        else:
            workbook = Workbook()  # Создание книги и листа Excel
            sheet = workbook.active
            # Напишите строку заголовка
            header = ['ID заявки', 'Дата', 'Номер гостя', 'Номер стола', 'Ф.И.О. гостя', 'Время', 'Комментарий']
            sheet.append(header)

            # Создаем объект PatternFill для изменения цвета
            bad_record_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

            for row in rows:
                if row[7] == 'bad':
                    row_data = list(row)[:7]  # Берем только первые 7 элементов
                    sheet.append(row_data)
                    for cell in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row, min_col=1, max_col=7):
                        for col_idx in range(1, len(cell) + 1):
                            cell[col_idx - 1].fill = bad_record_fill
                else:
                    sheet.append(row[:7])

            # Сохраните файл Excel
            file_name = f"отчет_НА_ЗАВТРА_{(datetime.date.today() + datetime.timedelta(days=1)).strftime('%Y-%m')}.xlsx"
            workbook.save(file_name)
            # Отправить файл Excel пользователю
            with open(file_name, 'rb') as file:
                await bot.send_document(callback_query.from_user.id, file)
            os.remove(file_name)  # Удаление файла
        conn.close()
    except Exception as e:
        await bot.send_message(callback_query.from_user.id, "Произошла ошибка при получении отчета на завтра.")
        print(e)


def apply_handler():
    """Регистрируем обработчики для (Заявки)"""
    dp.register_callback_query_handler(application_record_handler, text='apply')
    dp.register_callback_query_handler(send_application_handler, text='send_application')
    dp.register_callback_query_handler(edit_application_handler, text='edit_application')
    dp.register_callback_query_handler(send_report_for_today, text='report_for_today')
    dp.register_callback_query_handler(send_report_for_tomorrow, text='report_for_tomorrow')
